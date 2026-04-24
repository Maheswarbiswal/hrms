from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.utils.timezone import localtime
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Attendance
from hr.models import Employee
from django.http import HttpResponse
from datetime import datetime, date, time, timedelta
from calendar import monthrange
import pandas as pd
from django.utils.timezone import make_aware
import hashlib
from django.urls import reverse
SAFE_TIME = make_aware(datetime(1970, 1, 1, 0, 0))
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from leave.models import Holiday


# -------------------------------
# Custom Decorators
# -------------------------------

def login_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('user_authenticated'):
            messages.error(request, 'Please login to access this page.')
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


def role_required(allowed_roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            user_role = request.session.get('user_role')
            if not user_role or user_role not in allowed_roles:
                messages.error(request, 'You do not have permission to access this page.')
                return redirect('access_denied')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# -------------------------------
# Attendance Dashboard
# -------------------------------

@login_required
def attendance_dashboard(request):
    user_id = request.session.get('user_id')
    user_role = request.session.get('user_role')
    
    if user_role == 'ADMIN':
        messages.info(request, 'Admins can only view attendance.')
    
    employee = Employee.objects.get(id=user_id)
    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(employee=employee, date=today).first()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'check_in':
            if today_attendance:
                messages.warning(request, 'You have already checked in today.')
            else:
                Attendance.objects.create(
                    employee=employee,
                    date=today,
                    check_in=timezone.now()
                )
                messages.success(request, 'Check-in successful!')
                return redirect('attendance:dashboard')
        
        elif action == 'check_out':
            if not today_attendance:
                messages.error(request, 'You need to check in first.')
            elif today_attendance.check_out:
                messages.warning(request, 'You have already checked out today.')
            else:
                today_attendance.check_out = timezone.now()
                today_attendance.save()
                messages.success(request, 'Check-out successful!')
                return redirect('attendance:dashboard')
                
    
    context = {
        'today_attendance': today_attendance,
        'employee': employee,
    }
    return render(request, 'attendance/dashboard.html', context)


# -------------------------------
# View All Attendance (Employee)
# -------------------------------

@login_required
def all_attendance(request):
    user_id = request.session.get('user_id')
    user_role = request.session.get('user_role')

    employee = Employee.objects.get(id=user_id)
    attendance_records = Attendance.objects.filter(employee=employee).order_by('-date')

    month_filter = request.GET.get('month', '')
    today = date.today()

    if month_filter:
        year, month = map(int, month_filter.split('-'))
        start_date = date(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = date(year, month, last_day)

        if year == today.year and month == today.month:
            end_date = today
        elif date(year, month, last_day) > today:
            end_date = today
    else:
        end_date = today
        start_date = end_date - timedelta(days=30)

    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
        if (start_date + timedelta(days=i)) <= today
    ]

    attendance_dict = {att.date: att for att in attendance_records}
    full_attendance_list = []

    office_start_time = time(9, 30)

    late_arrivals_count = 0
    halfday_count = 0 
    total_extra_minutes = 0 

    for d in reversed(all_dates):
        record = attendance_dict.get(d)

        is_holiday = Holiday.objects.filter(
            date=d,
            region__name__iexact=employee.location
        ).exists()

        if record:
            record.duration_display = "-"
            record.extra_hours_display = "-"
            record.day_status = "Absent"

            record.punctuality = "—"

            # ===============================
            # EXISTING LOGIC (FIXED ONLY STATUS)
            # ===============================
            if record.check_in and record.check_out:
                diff = record.check_out - record.check_in
                total_minutes = diff.total_seconds() / 60
                hours = int(total_minutes // 60)
                minutes = int(total_minutes % 60)
                record.duration_display = f"{hours}h {minutes}m" if hours or minutes else "0 minutes"

                worked_hours = total_minutes / 60

                # ✅ FIXED HERE
                if worked_hours < 2:
                    record.day_status = "LOP"
                elif worked_hours < 5:
                    record.day_status = "Half Day"
                    if record.date.weekday() != 6: # Not Sunday
                        halfday_count += 1  
                else:
                    record.day_status = "Present"

                weekday = record.date.weekday()
                if weekday == 5:
                    if employee.location.strip().lower() == 'bhubaneswar':
                        standard_hours = 6  
                    else:
                        standard_hours = 4  
                else:
                    standard_hours = 9

                extra_minutes = int((worked_hours - standard_hours) * 60)

                if extra_minutes > 0:
                    total_extra_minutes += extra_minutes

                abs_minutes = abs(extra_minutes)
                eh_hours = abs_minutes // 60
                eh_mins = abs_minutes % 60

                if extra_minutes > 0:
                    record.extra_hours_display = f"+{eh_hours}h {eh_mins}m"
                elif extra_minutes < 0:
                    record.extra_hours_display = f"-{eh_hours}h {eh_mins}m"
                else:
                    record.extra_hours_display = "0h 0m"

            elif record.check_in and not record.check_out:
                record.duration_display = "In Progress"
                record.extra_hours_display = "-"

                # ❌ REMOVED WRONG LOGIC (datetime.now)
                # ✅ FIXED
                record.day_status = "Half Day"
                if record.date.weekday() != 6: # Not Sunday
                    halfday_count += 1  

            # ===============================
            # HOLIDAY / SUNDAY (KEEPED)
            # ===============================
            if not record.check_in:
                if is_holiday:
                    record.day_status = "Holiday"
                elif d.weekday() == 6:
                    record.day_status = "Sunday"

            # ===============================
            # PUNCTUALITY (UNCHANGED)
            # ===============================
            if record.check_in:
                local_checkin = record.check_in.astimezone().time()
                if local_checkin <= office_start_time:
                    record.punctuality = "On Time"
                else:
                    record.punctuality = "Late"
                    late_arrivals_count += 1 

            full_attendance_list.append(record)

        else:
            fake_record = Attendance(
                employee=employee,
                date=d,
                check_in=None,
                check_out=None,
            )
            fake_record.duration_display = "-"
            fake_record.extra_hours_display = "-"

            if is_holiday:
                fake_record.day_status = "Holiday"
            elif d.weekday() == 6:
                fake_record.day_status = "Sunday"
            else:
                fake_record.day_status = "Absent"

            full_attendance_list.append(fake_record)

    full_attendance_list.sort(key=lambda x: x.date, reverse=True)
    total_extra_hours = total_extra_minutes // 60
    total_extra_mins = total_extra_minutes % 60
    total_extra_display = f"{total_extra_hours}h {total_extra_mins}m"

    if not month_filter:
        month_filter = today.strftime("%Y-%m")

    context = {
        'attendances': full_attendance_list,
        'employee': employee,
        'selected_month': month_filter,
        'today': today,
        'late_arrivals_count': late_arrivals_count,
        'halfday_count': halfday_count,
        'total_extra_display': total_extra_display,
    }

    return render(request, 'attendance/all_attendance.html', context)

# -------------------------------
# Admin / HR Attendance Report
# -------------------------------

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN','BRANCH MANAGER'])
def attendance_report(request):
    search_query = request.GET.get('search', '')
    branch = request.GET.get('branch', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status_filter', '')

    today = date.today()
    office_start_time = time(9, 30)

    try:
        start_date = datetime.strptime(date_from, '%d/%m/%Y').date() if date_from else today
    except ValueError:
        start_date = today

    try:
        end_date = datetime.strptime(date_to, '%d/%m/%Y').date() if date_to else today
    except ValueError:
        end_date = today

    if end_date > today:
        end_date = today
    if end_date < start_date:
        end_date = start_date
    # ✅ Get current user details
    user_role = request.session.get('user_role')
    user_email = request.session.get('user_email')

    # ✅ Employees - Role-based filtering
    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        # Show all employees for admin roles
        employees = Employee.objects.all().order_by('first_name', 'last_name')
        filter_info = "Showing all employees"
    
    elif user_role == 'BRANCH MANAGER':
        try:
            # Get the current branch manager's employee record
            current_branch_manager = Employee.objects.get(email=user_email)
            
            # Get the branch manager's location
            branch_manager_location = current_branch_manager.location
            
            if branch_manager_location:
                # Show all employees with the same location
                employees = Employee.objects.filter(
                    location__iexact=branch_manager_location
                ).order_by('first_name', 'last_name')
                filter_info = f"Showing employees from {branch_manager_location} location"
            else:
                employees = Employee.objects.none()
                filter_info = "Branch manager location not set"
                
        except Employee.DoesNotExist:
            employees = Employee.objects.none()
            filter_info = "Branch manager profile not found"
    
    else:
        employees = Employee.objects.none()
        filter_info = "No access to attendance report"

     # ✅ Fixed branch filter (for CharField location) - only if user is admin/HR
    if branch and branch != 'All Branches' and user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        employees = employees.filter(location__iexact=branch)

    # ✅ For BRANCH MANAGER, always filter by their location regardless of branch filter
    elif user_role == 'BRANCH MANAGER':
        try:
            current_branch_manager = Employee.objects.get(email=user_email)
            if current_branch_manager.location:
                employees = employees.filter(location__iexact=current_branch_manager.location)
        except Employee.DoesNotExist:
            pass

    if search_query:
        names = search_query.strip().split()
        if len(names) == 1:
            employees = employees.filter(
                Q(first_name__icontains=names[0]) |
                Q(last_name__icontains=names[0]) |
                Q(employee_id__icontains=names[0])
            )
        else:
            first_part = names[0]
            last_part = " ".join(names[1:])
            employees = employees.filter(
                Q(first_name__icontains=first_part, last_name__icontains=last_part)
                | Q(employee_id__icontains=search_query)
            )

    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    attendance_data = []

    for emp in employees:
        for day in all_dates:
            is_holiday = Holiday.objects.filter(
                date=day,
                region__name__iexact=emp.location
            ).exists()
            att = Attendance.objects.filter(employee=emp, date=day).first()

            record = {
                'employee_pk': emp.id,
                'employee_id': emp.employee_id,
                'employee_name': f"{emp.first_name} {emp.last_name}",
                'branch': getattr(emp, 'location', '—'),
                'date': day,
                'check_in': "—",
                'check_out': "—",
                'checkin_address': "—",
                'checkout_address': "—",
                'status': 'Absent',
                'duration_display': '-',
                'is_late': False,
            }
           
            
            if att:
                if att.check_in:
                    ci = localtime(att.check_in)
                    record['check_in'] = ci.strftime("%I:%M %p")
                    record['checkin_address'] = att.checkin_address or "Location not available"
                    if ci.time() > office_start_time:
                        record['is_late'] = True

                if att.check_out:
                    co = localtime(att.check_out)
                    record['check_out'] = co.strftime("%I:%M %p")
                    record['checkout_address'] = att.checkout_address or "Location not available"

                #  UPDATED HALF-DAY / LOP LOGIC
                if att.check_in and att.check_out:
                    diff = att.check_out - att.check_in
                    total_minutes = diff.total_seconds() / 60
                    worked_hours = total_minutes / 60

                    if worked_hours < 2:
                        record['status'] = "LOP"
                    elif worked_hours < 5:
                        if day.weekday() == 5:
                            record['status'] = "Present"
                        else:
                            record['status'] = "Half Day"
                    else:
                        record['status'] = "Present"

                    hours = int(total_minutes // 60)
                    minutes = int(total_minutes % 60)
                    record['duration_display'] = f"{hours}h {minutes}m"

                elif att.check_in and not att.check_out:
                    # Check-in only = treat just like personal page
                    if day.weekday() == 5:
                        record['status'] = "Present"
                    else:
                        record['status'] = "Half Day"

                    record['duration_display'] = "In Progress"
                # ✅ HOLIDAY (ONLY IF NO ATTENDANCE)
                elif is_holiday:
                    record['status'] = "Holiday"

                # ✅ SUNDAY
                elif day.weekday() == 6:
                    record['status'] = "Sunday"

                # ✅ DEFAULT
                else:
                    record['status'] = "Absent"    

            else:
                if is_holiday:
                    record['status'] = "Holiday"
                elif day.weekday() == 6:
                    record['status'] = "Sunday"
                else:
                    record['status'] = "Absent"

            attendance_data.append(record)

    if status_filter:
        attendance_data = [a for a in attendance_data if a['status'] == status_filter]

    attendance_data.sort(key=lambda x: x['employee_name'].lower())
    attendance_records = attendance_data

     # ✅ Dynamic branches (from Employee table) - only show branches accessible to user
    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        branches = list(Employee.objects.values_list('location', flat=True).distinct())
    elif user_role == 'BRANCH MANAGER':
        try:
            current_branch_manager = Employee.objects.get(email=user_email)
            if current_branch_manager.location:
                branches = [current_branch_manager.location]
            else:
                branches = []
        except Employee.DoesNotExist:
            branches = []
    else:
        branches = []

    context = {
        'attendances': attendance_records,
        'branches': branches,
        'search_query': search_query,
        'selected_branch': branch,
        'date_from': date_from,
        'date_to': date_to,
        'today': today,
        'status_filter': status_filter,
    }

    return render(request, 'attendance/report.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def download_admin_attendance_report(request):
    """Download Excel with exactly the same filtered data shown in the dashboard."""
    search_query = request.GET.get('search', '')
    branch = request.GET.get('branch', '')
    department = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status_filter', '')

    today = date.today()
    office_start_time = time(9, 30)

    # Parse date range
    try:
        start_date = datetime.strptime(date_from, '%d/%m/%Y').date() if date_from else None
    except ValueError:
        start_date = None

    try:
        end_date = datetime.strptime(date_to, '%d/%m/%Y').date() if date_to else None
    except ValueError:
        end_date = None

    if not start_date and not end_date:
        start_date = end_date = today
    elif start_date and not end_date:
        end_date = start_date
    elif end_date and not start_date:
        start_date = end_date

    if end_date > today:
        end_date = today

      # ✅ Get current user details for role-based filtering
    user_role = request.session.get('user_role')
    user_email = request.session.get('user_email')

    # ✅ Employees - Role-based filtering (same as attendance_report)
    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        employees = Employee.objects.all().order_by('first_name', 'last_name')
    
    elif user_role == 'BRANCH MANAGER':
        try:
            current_branch_manager = Employee.objects.get(email=user_email)
            branch_manager_location = current_branch_manager.location
            
            if branch_manager_location:
                employees = Employee.objects.filter(
                    location__iexact=branch_manager_location
                ).order_by('first_name', 'last_name')
            else:
                employees = Employee.objects.none()
        except Employee.DoesNotExist:
            employees = Employee.objects.none()
    
    else:
        employees = Employee.objects.none()

     # ✅ Apply filters only for admin roles
    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        if department:
            employees = employees.filter(department=department)

        if branch and branch != 'All Branches':
            employees = employees.filter(location__iexact=branch)

    # ✅ For BRANCH MANAGER, always filter by their location
    elif user_role == 'BRANCH MANAGER':
        try:
            current_branch_manager = Employee.objects.get(email=user_email)
            if current_branch_manager.location:
                employees = employees.filter(location__iexact=current_branch_manager.location)
        except Employee.DoesNotExist:
            pass

    # ==============================
    # FIXED SEARCH FILTER (Copied from dashboard)
    # ==============================
    if search_query:
        names = search_query.strip().split()

        # One word search
        if len(names) == 1:
            employees = employees.filter(
                Q(first_name__icontains=names[0]) |
                Q(last_name__icontains=names[0]) |
                Q(employee_id__icontains=names[0])
            )
        else:
            # Two or more words → first + last name search
            first_part = names[0]
            last_part = " ".join(names[1:])
            employees = employees.filter(
                Q(first_name__icontains=first_part, last_name__icontains=last_part)
                | Q(employee_id__icontains=search_query)
            )

    # ==============================

    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    attendance_data = []

    for emp in employees:
        for day in all_dates:
            att = Attendance.objects.filter(employee=emp, date=day).first()
            is_holiday = Holiday.objects.filter(
                date=day,
                region__name__iexact=emp.location
            ).exists()

            record = {
                'employee_id': emp.employee_id,
                'employee_name': f"{emp.first_name} {emp.last_name}",
                'branch': getattr(emp, 'location', '—'),
                'date': day.strftime("%b %d, %Y"),
                'check_in': "—",
                'check_out': "—",
                'status': 'Absent',
                'duration_display': '-'
            }

            # ✅ SUNDAY (NO RECORD)
            if day.weekday() == 6 and not att:
                record['status'] = "Sunday"
                attendance_data.append(record)
                continue

            # =========================
            # ✅ MAIN LOGIC
            # =========================
            if att:

                # TIME
                if att.check_in:
                    record['check_in'] = localtime(att.check_in).strftime("%I:%M %p")
                if att.check_out:
                    record['check_out'] = localtime(att.check_out).strftime("%I:%M %p")

                # 🔥 IF CHECK-IN EXISTS → NORMAL STATUS (override holiday)
                if att.check_in:

                    if att.check_in and att.check_out:
                        diff = att.check_out - att.check_in
                        total_minutes = diff.total_seconds() / 60
                        worked_hours = total_minutes / 60

                        if worked_hours < 2:
                            record['status'] = "LOP"
                        elif worked_hours < 5:
                            if day.weekday() == 5:
                                record['status'] = "Present"
                            else:
                                record['status'] = "Half Day"
                        else:
                            record['status'] = "Present"

                        hours = int(total_minutes // 60)
                        minutes = int(total_minutes % 60)
                        record['duration_display'] = f"{hours}h {minutes}m"

                    elif att.check_in and not att.check_out:
                        if day.weekday() == 5:
                            record['status'] = "Present"
                        else:
                            record['status'] = "Half Day"
                        record['duration_display'] = "In Progress"

                # 🔥 NO CHECK-IN → CHECK HOLIDAY
                else:
                    if is_holiday:
                        record['status'] = "Holiday"   # or "HL"
                    else:
                        record['status'] = "Absent"

            else:
                # 🔥 NO ATTENDANCE
                if is_holiday:
                    record['status'] = "Holiday"   # or "HL"
                else:
                    record['status'] = "Absent"

            attendance_data.append(record)
    # Filters
    if status_filter:
        attendance_data = [a for a in attendance_data if a['status'] == status_filter]

    attendance_data.sort(key=lambda x: x['employee_name'].lower())

    if not attendance_data:
        response = HttpResponse(
            "No attendance data found for the selected date range and filters.",
            content_type="text/plain"
        )
        response['Content-Disposition'] = 'attachment; filename=\"Empty_Attendance_Report.txt\"'
        return response

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    align_center = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = [
        "Employee ID", "Employee Name", "Branch",
        "Date", "Check-In", "Check-Out", "Status", "Duration"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_style

    for record in attendance_data:
        ws.append([
            record['employee_id'],
            record['employee_name'],
            record['branch'],
            record['date'],
            record['check_in'],
            record['check_out'],
            record['status'],
            record['duration_display'],
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = align_center
            cell.border = border_style

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    filename = (
        f"Attendance_Report_{start_date.strftime('%b_%d_%Y')}"
        f"_to_{end_date.strftime('%b_%d_%Y')}.xlsx"
    )
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response


# -------------------------------
# Sync Attendance from Biometric Machine
# -------------------------------
from attendance.services import BiometricSyncService

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def sync_biometric_attendance(request):
    if request.method == 'POST':
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        
        if not date_from or not date_to:
            # Default to current month or past 7 days if you prefer
            today = timezone.now()
            start = today.replace(day=1).strftime('%Y-%m-%d 00:00')
            end = today.strftime('%Y-%m-%d 23:59')
        else:
            try:
                # Assuming date_from and date_to come in DD/MM/YYYY or YYYY-MM-DD
                # Front end report page returns DD/MM/YYYY usually
                if '/' in date_from:
                    d_from = datetime.strptime(date_from, '%d/%m/%Y')
                    d_to = datetime.strptime(date_to, '%d/%m/%Y')
                else:
                    d_from = datetime.strptime(date_from, '%Y-%m-%d')
                    d_to = datetime.strptime(date_to, '%Y-%m-%d')
                
                start = d_from.strftime('%Y-%m-%d 00:00')
                end = d_to.strftime('%Y-%m-%d 23:59')
            except Exception as e:
                messages.error(request, f"Invalid date format: {e}")
                return redirect('attendance:report')

        result = BiometricSyncService.sync_attendance(start, end)
        
        if result['status'] == 'success':
            messages.success(request, result['message'])
            # if result.get('missing_biometric_ids'):
                # messages.warning(request, f"Unmapped Biometric IDs found: {', '.join(result['missing_biometric_ids'])}. Please map them in Employee settings.")
        else:
            messages.error(request, result['message'])
            
    return redirect('attendance:report')

# -------------------------------
# Generate Attendance Report PDF
# -------------------------------

@login_required
def download_attendance_report_excel(request):
    """Download Excel attendance report for a single employee (monthly view)."""
    user_id = request.session.get('user_id')
    employee = Employee.objects.get(id=user_id)

    # ✅ Get selected month
    month_filter = request.GET.get('month', '')
    today = date.today()

    if month_filter:
        year, month = map(int, month_filter.split('-'))
    else:
        year, month = today.year, today.month

    # ✅ Month range
    start_date = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_date = date(year, month, last_day)

    # ✅ Restrict to current day for live month
    if year == today.year and month == today.month:
        end_date = today
    elif end_date > today:
        end_date = today

    # ✅ Get existing attendance data
    attendance_records = Attendance.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    )
    attendance_dict = {a.date: a for a in attendance_records}

    # ✅ Generate list of all days
    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    # ✅ Prepare full attendance list
    full_attendance_list = []
    for d in all_dates:

        is_holiday = Holiday.objects.filter(
            date=d,
            region__name__iexact=employee.location
        ).exists()

        if d in attendance_dict:
            record = attendance_dict[d]

            check_in = localtime(record.check_in).strftime("%I:%M %p") if record.check_in else "-"
            check_out = localtime(record.check_out).strftime("%I:%M %p") if record.check_out else "-"

            # =========================
            # ✅ CORRECT LOGIC (MATCH PORTAL)
            # =========================

            if record.check_in or record.check_out:

                if record.check_in and record.check_out:
                    diff = record.check_out - record.check_in
                    total_minutes = diff.total_seconds() / 60
                    worked_hours = total_minutes / 60

                    hours = int(total_minutes // 60)
                    minutes = int(total_minutes % 60)
                    duration = f"{hours}h {minutes}m"

                    if worked_hours < 2:
                        status = "LOP"
                    elif worked_hours < 5:
                        status = "Half Day"
                    else:
                        status = "Present"

                elif record.check_in and not record.check_out:
                    status = "Half Day"
                    duration = "In Progress"

                else:
                    status = "Absent"
                    duration = "-"

            else:
                if d.weekday() == 6:
                    status = "Sunday"   # Week Off
                elif is_holiday:
                    status = "Holiday"
                else:
                    status = "Absent"

                duration = "-"

        else:
            check_in = "-"
            check_out = "-"

            if d.weekday() == 6:
                status = "Sunday"
            elif is_holiday:
                status = "Holiday"
            else:
                status = "Absent"

            duration = "-"

        full_attendance_list.append({
            'date': d.strftime("%b %d, %Y"),
            'check_in': check_in,
            'check_out': check_out,
            'status': status,
            'duration': duration,
        })

    # ✅ Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    align_center = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = ["Date", "Check-In", "Check-Out", "Status", "Duration"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_style

    for record in full_attendance_list:
        ws.append([
            record['date'],
            record['check_in'],
            record['check_out'],
            record['status'],
            record['duration']
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = align_center
            cell.border = border_style

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    month_name = start_date.strftime("%B")
    filename = f"{employee.first_name}_{month_name}_{year}_Attendance_Report.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def upload_admin_attendance_excel(request):

    if request.method != "POST":
        return redirect("attendance:report")

    excel_file = request.FILES.get("attendance_file")
    current_branch = request.POST.get("current_branch", "")

    if not excel_file:
        messages.error(request, "Please upload a valid Excel file.")
        return redirect(f"{reverse('attendance:report')}?branch={current_branch}")

    # ============================
    # LOAD EXCEL
    # ============================
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
    except:
        messages.error(request, "Invalid Excel file.")
        return redirect(f"{reverse('attendance:report')}?branch={current_branch}")

    new_updates = 0

    # ============================
    # READ ROWS
    # ============================
    for row in sheet.iter_rows(min_row=2, values_only=True):

        employee_id = str(row[0]).strip() if row[0] else None
        if not employee_id:
            continue

        # ==========================================================
        # ✔ BRANCH VALIDATION – STOP WRONG LOCATION UPLOAD
        # ==========================================================
        try:
            employee = Employee.objects.get(employee_id__iexact=employee_id)
        except:
            continue

        # FIXED HERE: employee.location is correct field
        if current_branch and employee.location != current_branch:
            messages.error(
                request,
                f"❌ Excel contains employees from **{employee.location}**, "
                f"but selected location is **{current_branch}**."
            )
            return redirect(f"{reverse('attendance:report')}?branch={current_branch}")

        # ==========================================================
        # PROCESS DATE/TIME
        # ==========================================================
        excel_date = row[3]
        check_in_val = row[4]
        check_out_val = row[5]

        if isinstance(excel_date, datetime):
            date_obj = excel_date.date()
        else:
            try:
                date_obj = datetime.strptime(str(excel_date), "%b %d, %Y").date()
            except:
                continue

        def parse_time(val):
            if val in [None, "", "-", "—"]:
                return None
            if isinstance(val, datetime):
                return val.time()

            val = str(val).replace(":", ":").strip().upper()
            fmts = ["%I:%M %p", "%I:%M%p", "%I %p", "%H:%M"]
            for f in fmts:
                try:
                    return datetime.strptime(val, f).time()
                except:
                    pass
            return None

        in_time = parse_time(check_in_val)
        out_time = parse_time(check_out_val)

        excel_in = make_aware(datetime.combine(date_obj, in_time)) if in_time else None
        excel_out = make_aware(datetime.combine(date_obj, out_time)) if out_time else None

        existing = Attendance.objects.filter(employee=employee, date=date_obj).first()

        if existing:
            db_in = existing.check_in
            db_out = existing.check_out

            if db_in == excel_in and db_out == excel_out:
                continue

            existing.check_in = excel_in
            existing.check_out = excel_out
            existing.save()
            new_updates += 1

        else:
            Attendance.objects.create(
                employee=employee,
                date=date_obj,
                check_in=excel_in,
                check_out=excel_out
            )
            new_updates += 1

    # ==========================================================
    # NO NEW UPDATES
    # ==========================================================
    if new_updates == 0:
        messages.warning(
            request,
            "⚠ This Excel file contains no new attendance updates. It was already uploaded before."
        )
        return redirect(f"{reverse('attendance:report')}?branch={current_branch}")

    messages.success(request, f"{new_updates} attendance records updated successfully.")
    return redirect(f"{reverse('attendance:report')}?branch={current_branch}")

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN', 'BRANCH MANAGER'])
def attendance_details(request):
    month_filter = request.GET.get('month', '')
    branch = request.GET.get('branch', '')
    user_role = request.session.get('user_role')
    user_email = request.session.get('user_email')
    
    today = date.today()
    if not month_filter:
        month_filter = today.strftime('%Y-%m')
    
    try:
        year, month = map(int, month_filter.split('-'))
    except ValueError:
        year, month = today.year, today.month
        month_filter = today.strftime('%Y-%m')
        
    start_date = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_date = date(year, month, last_day)

    all_dates = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    day_names = [d.strftime('%a') for d in all_dates]

    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        branches = list(Employee.objects.values_list('location', flat=True).distinct())
        employees = Employee.objects.all().order_by('first_name', 'last_name')
    elif user_role == 'BRANCH MANAGER':
        try:
            current_branch_manager = Employee.objects.get(email=user_email)
            bm_loc = current_branch_manager.location
            branches = [bm_loc] if bm_loc else []
            employees = Employee.objects.filter(location__iexact=bm_loc).order_by('first_name', 'last_name')
        except Employee.DoesNotExist:
            branches = []
            employees = Employee.objects.none()
    else:
        branches = []
        employees = Employee.objects.none()

    if branch and branch != 'All Branches' and user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        employees = employees.filter(location__iexact=branch)

    report_data = []

    for emp in employees:
        emp_data = {
            'employee': emp,
            'days': [],
            'total_days': len(all_dates),
            'absent': 0,
            'working_days': 0,
        }
        
        attendances = Attendance.objects.filter(employee=emp, date__range=[start_date, end_date])
        att_dict = {a.date: a for a in attendances}
        
        for d in all_dates:
            if d > today:
                status = '-'
                emp_data['days'].append(status)
                continue

            att = att_dict.get(d)
            is_holiday = Holiday.objects.filter(
                date=d,
                region__name__iexact=emp.location
            ).exists()

            # =========================
            # SAME LOGIC AS REPORT (FIXED)
            # =========================
            if att:
                if att.check_in and att.check_out:
                    diff = att.check_out - att.check_in
                    worked_hours = diff.total_seconds() / 3600

                    if worked_hours < 2:
                        status_full = "LOP"
                        emp_data['absent'] += 1
                    elif worked_hours < 5:
                        if d.weekday() == 5: # Saturday
                            status_full = "Present"
                            emp_data['working_days'] += 1
                        elif d.weekday() == 6: # Sunday
                            status_full = "Half Day"
                            emp_data['working_days'] += 1
                            # No addition to absent for Sunday
                        else:
                            status_full = "Half Day"
                            emp_data['working_days'] += 0.5
                            emp_data['absent'] += 0.5
                    else:
                        status_full = "Present"
                        emp_data['working_days'] += 1

                elif att.check_in and not att.check_out:
                    if d.weekday() == 5: # Saturday
                        status_full = "Present"
                        emp_data['working_days'] += 1
                    elif d.weekday() == 6: # Sunday
                        status_full = "Half Day"
                        emp_data['working_days'] += 1
                        # No addition to absent for Sunday
                    else:
                        status_full = "Half Day"
                        emp_data['working_days'] += 0.5
                        emp_data['absent'] += 0.5

                else:
                    if is_holiday:   # ✅ ADDED
                        status_full = "Holiday"
                        emp_data['working_days'] += 1
                    elif d.weekday() == 6:
                        status_full = "Sunday"
                        emp_data['working_days'] += 1
                    else:
                        status_full = "Absent"
                        emp_data['absent'] += 1

            else:
                if is_holiday:   # ✅ ADDED
                    status_full = "Holiday"
                    emp_data['working_days'] += 1
                elif d.weekday() == 6:
                    status_full = "Sunday"
                    emp_data['working_days'] += 1
                else:
                    status_full = "Absent"
                    emp_data['absent'] += 1

            # =========================
            # MAPPING (UNCHANGED)
            # =========================
            status_map = {
                "Present": "P",
                "Absent": "A",
                "Half Day": "HD",
                "Holiday": "HL",
                "LOP": "L",
                "Sunday": "S"
            }

            status = status_map.get(status_full, "A")

            emp_data['days'].append(status)
            
        report_data.append(emp_data)

    context = {
        'month': month_filter,
        'branches': branches,
        'selected_branch': branch,
        'day_names': day_names,
        'report_data': report_data
    }

    return render(request, 'attendance/attendance_details.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN', 'BRANCH MANAGER'])
def download_monthly_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    
    month_filter = request.GET.get('month', '')
    branch = request.GET.get('branch', '')
    user_role = request.session.get('user_role')
    user_email = request.session.get('user_email')
    
    today = date.today()
    if not month_filter:
        month_filter = today.strftime('%Y-%m')
        
    try:
        year, month = map(int, month_filter.split('-'))
    except ValueError:
        year, month = today.year, today.month
        
    # ✅ FULL MONTH (FIXED)
    start_date = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_date = date(year, month, last_day)

    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    # =========================
    # EMPLOYEE FILTER
    # =========================
    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        employees = Employee.objects.all().order_by('first_name', 'last_name')
    elif user_role == 'BRANCH MANAGER':
        try:
            bm_loc = Employee.objects.get(email=user_email).location
            employees = Employee.objects.filter(
                location__iexact=bm_loc
            ).order_by('first_name', 'last_name')
        except:
            employees = Employee.objects.none()
    else:
        employees = Employee.objects.none()

    if branch and branch != 'All Branches' and user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        employees = employees.filter(location__iexact=branch)

    # =========================
    # EXCEL SETUP
    # =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    align_center = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["Employee Name"] + [str(d.day) for d in all_dates] + ["Total Days", "Absent", "Working Days"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_style

    # =========================
    # DATA GENERATION
    # =========================
    for emp in employees:
        row = [f"{emp.first_name} {emp.last_name}"]

        attendances = Attendance.objects.filter(
            employee=emp,
            date__range=[start_date, end_date]
        )
        att_dict = {a.date: a for a in attendances}

        absent = 0
        working = 0

        for d in all_dates:

            # ✅ FUTURE DATE FIX
            if d > today:
                row.append('-')
                continue

            att = att_dict.get(d)

            is_holiday = Holiday.objects.filter(
                date=d,
                region__name__iexact=emp.location
            ).exists()

            # =========================
            # SAME LOGIC AS UI
            # =========================
            if att:
                if att.check_in and att.check_out:
                    diff = att.check_out - att.check_in
                    worked_hours = diff.total_seconds() / 3600

                    if worked_hours < 2:
                        status_full = "LOP"
                        absent += 1
                    elif worked_hours < 5:
                        if d.weekday() == 5:
                            status_full = "Present"
                            working += 1
                        else:
                            status_full = "Half Day"
                            working += 0.5
                            absent += 0.5
                    else:
                        status_full = "Present"
                        working += 1

                elif att.check_in and not att.check_out:
                    if d.weekday() == 5:
                        status_full = "Present"
                        working += 1
                    else:
                        status_full = "Half Day"
                        working += 0.5
                        absent += 0.5

                else:
                    if is_holiday:
                        status_full = "Holiday"
                        working += 1
                    elif d.weekday() == 6:
                        status_full = "Sunday"
                        working += 1
                    else:
                        status_full = "Absent"
                        absent += 1

            else:
                if is_holiday:
                    status_full = "Holiday"
                    working += 1
                elif d.weekday() == 6:
                    status_full = "Sunday"
                    working += 1
                else:
                    status_full = "Absent"
                    absent += 1

            status_map = {
                "Present": "P",
                "Absent": "A",
                "Half Day": "HD",
                "Holiday": "HL",
                "LOP": "L",
                "Sunday": "S"
            }

            row.append(status_map.get(status_full, "A"))

        row.extend([len(all_dates), absent, working])
        ws.append(row)

    # =========================
    # STYLING
    # =========================
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in r:
            cell.alignment = align_center
            cell.border = border_style

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    filename = f"Monthly_Attendance_{start_date.strftime('%b_%Y')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response